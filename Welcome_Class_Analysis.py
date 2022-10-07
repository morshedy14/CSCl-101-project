import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from Functions_Class_Analysis import *
from PyQt5.QtWidgets import *
from PyQt5 import uic
from matplotlib import pyplot as plt



class Ui_Wlcome_Class_Analysis(QWidget):
    def setupUi(self, Wlcome_Class_Analysis):
        Wlcome_Class_Analysis.setObjectName("Wlcome_Class_Analysis")
        Wlcome_Class_Analysis.resize(593, 373)
        self.centralwidget = QtWidgets.QWidget(Wlcome_Class_Analysis)
        self.centralwidget.setObjectName("centralwidget")
        self.label_Welcome_Class = QtWidgets.QLabel(self.centralwidget)
        self.label_Welcome_Class.setGeometry(QtCore.QRect(120, -20, 400, 121))
        font = QtGui.QFont()
        font.setPointSize(18)
        self.label_Welcome_Class.setFont(font)
        self.label_Welcome_Class.setObjectName("label_Welcome_Class")
        self.label_Type = QtWidgets.QLabel(self.centralwidget)
        self.label_Type.setGeometry(QtCore.QRect(110, 90, 500, 61))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label_Type.setFont(font)
        self.label_Type.setObjectName("label_Type")
        self.pushButton_Text_file = QtWidgets.QPushButton(self.centralwidget, clicked = lambda: self.gettext())
        self.pushButton_Text_file.setGeometry(QtCore.QRect(50, 200, 141, 101))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.pushButton_Text_file.setFont(font)
        self.pushButton_Text_file.setObjectName("pushButton_Text_file")
        self.pushButton_Excel_file = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_Excel_file.setGeometry(QtCore.QRect(220, 200, 141, 101))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.pushButton_Excel_file.setFont(font)
        self.pushButton_Excel_file.setObjectName("pushButton_Excel_file")
        self.pushButton_CSV_file = QtWidgets.QPushButton(self.centralwidget, clicked = lambda: self.getCSV())
        self.pushButton_CSV_file.setGeometry(QtCore.QRect(400, 200, 141, 101))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.pushButton_CSV_file.setFont(font)
        self.pushButton_CSV_file.setObjectName("pushButton_CSV_file")
        Wlcome_Class_Analysis.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Wlcome_Class_Analysis)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 593, 21))
        self.menubar.setObjectName("menubar")
        Wlcome_Class_Analysis.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Wlcome_Class_Analysis)
        self.statusbar.setObjectName("statusbar")
        Wlcome_Class_Analysis.setStatusBar(self.statusbar)

        self.retranslateUi(Wlcome_Class_Analysis)
        QtCore.QMetaObject.connectSlotsByName(Wlcome_Class_Analysis)

    def retranslateUi(self, Wlcome_Class_Analysis):
        _translate = QtCore.QCoreApplication.translate
        Wlcome_Class_Analysis.setWindowTitle(_translate("Wlcome_Class_Analysis", "MainWindow"))
        self.label_Welcome_Class.setText(_translate("Wlcome_Class_Analysis", "Welcome To Class Analysis"))
        self.label_Type.setText(_translate("Wlcome_Class_Analysis", "What type of files do you prefer to use?"))
        self.pushButton_Text_file.setText(_translate("Wlcome_Class_Analysis", "Text File"))
        self.pushButton_Excel_file.setText(_translate("Wlcome_Class_Analysis", "Excel File"))
        self.pushButton_CSV_file.setText(_translate("Wlcome_Class_Analysis", "CSV File"))

    def gettext(self):
        text, ok = QInputDialog.getText(self, 'from text file', 'Enter file path:')
        if ok:
            self.MT = self.data_txt(text)
            self.graph_plain_numbers_or_txt()

    def getCSV(self):
        text, ok = QInputDialog.getText(self, 'from csv file', 'Enter file path:')
        if ok:
            self.MT = self.data_txt(text)
            self.graph_CSV(text, "Grades")


    def graph_plain_numbers_or_txt(self):

        MT_sorted = sort(self.MT)  # To graph a histogram, numbers should be sorted and in ascending order
        Average = average(MT_sorted)  # Calculates the average of dataset
        Median = median(MT_sorted)  # Calculates the median of the dataset
        Bins = range(5, 105, 11)  # Specifying our custom bins (x-axis values)

        plt.style.use("ggplot")  # Uses a famous styling sheet called "ggplot"
        plt.hist(MT_sorted, bins=Bins, edgecolor="black")  # Graphs midterm grades versus number of students in a histogram
        plt.title("Grade Distribution", fontsize=25, font="Arial")  # Sets graph title to "Grade Distribution"

        plt.xlabel("Grades", fontsize=20)  # Sets "Grades" as the title for the x-axis
        plt.ylabel("No. Students", fontsize=20)  # Sets "No. Students" as the title for the y-axis
        plt.axvline(Median, color="pink", label="Median", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our median
        plt.axvline(Average, color="purple", label="Average", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our average
        plt.xticks(Bins)  # Displays our custom bin values instead of the default one (multiples of 10)

        plt.tight_layout()  # Positions axis titles correctly in place
        plt.grid(True, color="white", axis="y")  # Plots a parallel-x-axis gird
        plt.legend()  # Displays the graph legend at the top-right most of the screen

        plt.show(block=False)  # Displays our graph and does not halt the execution of the program
        plt.pause(6)  # Pauses the execution for a specific time interval to allow the user to comprehend the graph
        plt.savefig("Figure_Class_Grades.pdf")  # initially creates an empty file
        plt.savefig("Figure_Class_Grades.pdf")  # Saves the figure as pdf


#======================================================================

    def data_txt(self, file_path):
        # Function to extract data from a text file
        file = open(file_path, "r")
        MT = []
        for line in file:
            if line == "Grades\n": continue
            MT.append(float(line))
        return MT

    def data_csv(file_path, column_name):
        # Function to extract the data of the CSV file
        import pandas as pd  # Imports pandas, a well-known python library for data science/analysis
        data = pd.read_csv(file_path)  # Reading midterm grades from the specified .csv file

        MT = []
        for index in data.index:
            each_num = data[column_name][index]
            MT.append(each_num)
        return MT

    def graph_CSV(self, file_path, column_name):
        # Function to graph the data of students into a histogram
        from matplotlib import \
            pyplot as plt  # Imports matplotlib, a widely-known graphing and data visualization python library
        import pandas as pd  # Imports pandas, a well-known python library for data science/analysis

        data = pd.read_csv(file_path)  # Reading midterm grades from the specified .csv file
        MT_grades = data[
            column_name]  # Specifying the name of the column that we will graph (useful in case of many columns)

        MT_sorted = sort(MT_grades)  # To graph a histogram, numbers should be sorted and in ascending order
        Average = average(MT_sorted)  # Calculates the average of dataset
        Median = median(MT_sorted)  # Calculates the median of the dataset
        Bins = range(5, 105, 11)  # Specifying our custom bins (x-axis values)

        plt.style.use("dark_background")  # Sets the background color to dark instead of the default white
        plt.hist(MT_sorted, bins=Bins, color="navy",
                 edgecolor="white")  # Graphs midterm grades versus number of students in a histogram
        plt.title("Grade Distribution", color="white", fontsize=25,
                  font="Arial")  # Sets graph title to "Grade Distribution"

        plt.xlabel("Grades", color="white", fontsize=20)  # Sets "Grades" as the title for the x-axis
        plt.ylabel("No. Students", color="white", fontsize=20)  # Sets "No. Students" as the title for the y-axis
        plt.axvline(Median, color="red", label="Median",
                    linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our median
        plt.axvline(Average, color="orange", label="Average",
                    linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our average
        plt.xticks(Bins)  # Displays our custom bin values instead of the default one (multiples of 10)

        plt.tight_layout()  # Positions axis titles correctly in place
        plt.grid(True, color="dimgray", axis="y")  # Plots a parallel-x-axis gird
        plt.legend()  # Displays the graph legend at the top-right most of the screen

        plt.show(block=False)  # Displays our graph and does not halt the execution of the program
        plt.pause(6)  # Pauses the execution for a specific time interval to allow the user to comprehend the graph
        plt.savefig("Figure_Class_Grades.pdf")  # initially creates an empty file
        plt.savefig("Figure_Class_Grades.pdf")  # Saves the figure as pdf

#==============================================================================================================================
    def open_excel(self, file_path):
        # Function to open and read data in Excel file
        from openpyxl import load_workbook
        workbook = load_workbook(file_path, read_only=True)
        return workbook


    def correct_sheet_name(self, sheet_name, file_path):
        # Function to check whether the entered sheet name actually exists with the Excel file
        workbook = open_excel(file_path)
        sheet_names = workbook.sheetnames
        while sheet_name not in sheet_names:
            sheet_name = input("Error! this sheet name does not exist, Please enter sheet name: ")
        return sheet_name


    def data_xlsx(self, file_path, sheet_name, column_name):
        # Function to extract the data of the Excel file
        import pandas as pd  # Imports pandas, a well-known python library for data science/analysis
        data = pd.read_excel(file_path, sheet_name=sheet_name)  # Reading midterm grades from the specified .xlsx file

        MT = []
        for index in data.index:
            each_num = data[column_name][index]
            MT.append(each_num)
        return MT


    def graph_xlsx(self, file_path, sheet_name, column_name):
        # Function to graph the data of students into a histogram
        from matplotlib import pyplot as plt  # Imports matplotlib, a widely-known graphing and data visualization python library
        import pandas as pd  # Imports pandas, a well-known python library for data science/analysis
        import mplcyberpunk  # Imports a cool styling sheet cyberpunk-like



        data = pd.read_excel(file_path, sheet_name=sheet_name)  # Reading midterm grades from a .xlsx file
        MT_grades = data[column_name]  # Specifying the name of the column that we will graph (useful in case of many columns)

        MT_sorted = sort(MT_grades)  # To graph a histogram, numbers should be sorted and in ascending order
        Average = average(MT_sorted)  # Calculates the average of dataset
        Median = median(MT_sorted)  # Calculates the median of the dataset
        Bins = range(5, 105, 11)  # Specifying our custom bins (x-axis values)

        plt.style.use("cyberpunk")  # Sets the background color to dark instead of the default white
        plt.hist(MT_sorted, bins=Bins, edgecolor="black")  # Graphs midterm grades versus number of students in a histogram
        plt.title("Grade Distribution", fontsize=25, font="Arial")  # Sets graph title to "Grade Distribution"\

        plt.xlabel("Grades", fontsize=20)  # Sets "Grades" as the title for the x-axis
        plt.ylabel("No. Students", fontsize=20)  # Sets "No. Students" as the title for the y-axis
        plt.axvline(Median, color="red", label="Median", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our median
        plt.axvline(Average, color="orange", label="Average", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our average
        plt.xticks(Bins)  # Displays our custom bin values instead of the default one (multiples of 10)

        plt.tight_layout()  # Positions axis titles correctly in place
        plt.grid(True, axis="y")  # Plots a parallel-x-axis gird
        plt.legend()  # Displays the graph legend at the top-right most of the screen

        plt.show(block=False)  # Displays our graph and does not halt the execution of the program
        plt.pause(5)  # Pauses the execution for a specific time interval to allow the user to comprehend the graph
        plt.savefig("Figure_Class_Grades.pdf")  # initially creates an empty file
        plt.savefig("Figure_Class_Grades.pdf")  # Saves the figure as pdf




if __name__ == "__main__":
    import sys
    print("HERE")
    app = QtWidgets.QApplication(sys.argv)
    Wlcome_Class_Analysis = QtWidgets.QMainWindow()
    ui = Ui_Wlcome_Class_Analysis()
    ui.setupUi(Wlcome_Class_Analysis)
    Wlcome_Class_Analysis.show()
    sys.exit(app.exec_())
