

def Write(file_name):
    # Function to write the output into a new file
    Write = open(file_name, "w")
    return Write


def correct_MT(MT):
    # Function to check if the user entered correct midterm grades
    while any(float(i) < 0 for i in MT) or any(float(i) > 105 for i in MT):
        MT = input("Error! please enter correct midterm grades: ").split()
    return MT


def correct_grade_type(choose):
    # Function to check if the user entered a correct choice for the dataset of grades (plain numbers or file)
    while choose not in ["1", "2", "3", "4"]:
        choose = input("Error! Do you have the dataset as plain numbers (directly into terminal) \"1\" "
                       " a CSV file, \"2\" an Excel file, \"3\" or a text file \"4\": ")
    return choose


def correct_YesNo(response):
    # Function to check if the user entered a correct choice for the dataset of grades (plain numbers or file)
    while response != "N" and response != "Y":
        response = input("Error! Would you like to keep a copy of these statistics, yes \"Y\" or no \"N\": ")
    return response


def correct_txt(file_name):
    # Function to check whether the user entered a correct .txt file
    while file_name.lower().endswith(".txt") == False:
        file_name = input("Error! Please enter the name of the text file: ")
    return file_name


def correct_file_type(file_type):
    # Function to check whether the user entered a correct file type
    while file_type != "1" and file_type != "2" and file_type != "3":
        file_type = input("Error! Is it a text file \"1\", an excel file \"2\", or a csv file \"3\": ")
    return file_type


def correct_xlsx(file_name):
    # Function to check whether the user entered a correct .xlsx file
    while file_name.lower().endswith(".xlsx") == False:
        file_name = input("Error! Please enter the name of the excel file: ")
    return file_name


def correct_csv(file_name):
    # Function to check whether the user entered a correct .csv file
    while file_name.lower().endswith(".csv") == False:
        file_name = input("Error! Please enter the name of the csv file: ")
    return file_name


def sort(MT):
    # Function to sort midterm grades in ascending order (or any other list input)
    for i in range(len(MT)):
        for j in range(i+1, len(MT)):
            if MT[i] > MT[j]:
                tmp1 = MT[i]
                MT[i] = MT[j]
                MT[j] = tmp1
    MT_sorted = MT
    return MT_sorted


def highest_lowest(MT):
    # Function to determine the highest and lowest values of the entered midterm grades after being sorted in ascending order
    sort(MT)
    Lowest_grade = MT[0]
    Highest_grade = MT[-1]

    return Highest_grade, Lowest_grade


def No_students_seg(MT):
    # Function to calculate the number of students in each segment of grades

    No_students = []
    c1 = c2 = c3 = c4 = c5 = c6 = c7 = c8 = c9 = c10 = 0
    MT_float(MT)

    for i in range(len(MT)):
        if MT[i] >= 0 and MT[i] <= 5:
            c1 = c1 + 1

        elif MT[i] > 5 and MT[i] <= 16:
            c2 = c2 + 1

        elif MT[i] > 16 and MT[i] <= 27:
            c3 = c3 + 1

        elif MT[i] > 27 and MT[i] <= 38:
            c4 = c4 + 1

        elif MT[i] > 38 and MT[i] <= 49:
            c5 = c5 + 1

        elif MT[i] > 49 and MT[i] <= 60:
            c6 = c6 + 1

        elif MT[i] > 60 and MT[i] <= 71:
            c7 = c7 + 1

        elif MT[i] > 71 and MT[i] <= 82:
            c8 = c8 + 1

        elif MT[i] > 82 and MT[i] <= 93:
            c9 = c9 + 1

        elif MT[i] > 93 and MT[i] <= 104:
            c10 = c10 + 1

    No_students.extend([c1, c2, c3, c4, c5, c6, c7, c8, c9, c10])  # Used .extend instead of .append to insert multiple values to a list
    return No_students


def graph_plain_numbers_or_txt(MT_grades):
    # Function to graph the data of students into a histogram (input as plain numbers, space-separated)
    from matplotlib import pyplot as plt  # Imports matplotlib, a widely-known graphing and data visualization python library

    MT_sorted = sort(MT_grades)  # To graph a histogram, numbers should be sorted and in ascending order
    Average = average(MT_sorted)  # Calculates the average of dataset
    Median = median(MT_sorted)  # Calculates the median of the dataset
    Bins = range(5, 105, 11)  # Specifying our custom bins (x-axis values)

    plt.style.use("ggplot")  # Uses a famous styling sheet called "ggplot"
    plt.hist(MT_sorted, bins=Bins, edgecolor="black")  # Graphs midterm grades versus number of students in a histogram
    plt.title("Grade Distribution", fontsize=25, font="Arial")  # Sets graph title to "Grade Distribution"

    plt.xlabel("Grades", fontsize=20)  # Sets "Grades" as the title for the x-axis
    plt.ylabel("No. Students", fontsize=20)  # Sets "No. Students" as the title for the y-axis
    plt.axvline(Median, color="pink", label="Median", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our median
    plt.axvline(Average, color="purple", label="Average", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our average
    plt.xticks(Bins)  # Displays our custom bin values instead of the default one (multiples of 10)

    plt.tight_layout()  # Positions axis titles correctly in place
    plt.grid(True, color="white", axis="y")  # Plots a parallel-x-axis gird
    plt.legend()  # Displays the graph legend at the top-right most of the screen

    plt.show(block=False)  # Displays our graph and does not halt the execution of the program
    plt.pause(6)  # Pauses the execution for a specific time interval to allow the user to comprehend the graph
    plt.savefig("Figure_Class_Grades.pdf")  # initially creates an empty file
    plt.savefig("Figure_Class_Grades.pdf")  # Saves the figure as pdf


def data_csv(file_path, column_name):
    # Function to extract the data of the CSV file
    import pandas as pd  # Imports pandas, a well-known python library for data science/analysis
    data = pd.read_csv(file_path)  # Reading midterm grades from the specified .csv file

    MT = []
    for index in data.index:
        each_num = data[column_name][index]
        MT.append(each_num)
    return MT


def graph_CSV(file_path, column_name):
    # Function to graph the data of students into a histogram
    from matplotlib import pyplot as plt  # Imports matplotlib, a widely-known graphing and data visualization python library
    import pandas as pd  # Imports pandas, a well-known python library for data science/analysis

    data = pd.read_csv(file_path)  # Reading midterm grades from the specified .csv file
    MT_grades = data[column_name]  # Specifying the name of the column that we will graph (useful in case of many columns)

    MT_sorted = sort(MT_grades)  # To graph a histogram, numbers should be sorted and in ascending order
    Average = average(MT_sorted)  # Calculates the average of dataset
    Median = median(MT_sorted)  # Calculates the median of the dataset
    Bins = range(5, 105, 11)  # Specifying our custom bins (x-axis values)

    plt.style.use("dark_background")  # Sets the background color to dark instead of the default white
    plt.hist(MT_sorted, bins=Bins, color="navy", edgecolor="white")  # Graphs midterm grades versus number of students in a histogram
    plt.title("Grade Distribution", color="white", fontsize=25, font="Arial")  # Sets graph title to "Grade Distribution"

    plt.xlabel("Grades", color="white", fontsize=20)  # Sets "Grades" as the title for the x-axis
    plt.ylabel("No. Students", color="white", fontsize=20)  # Sets "No. Students" as the title for the y-axis
    plt.axvline(Median, color="red", label="Median", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our median
    plt.axvline(Average, color="orange", label="Average", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our average
    plt.xticks(Bins)  # Displays our custom bin values instead of the default one (multiples of 10)

    plt.tight_layout()  # Positions axis titles correctly in place
    plt.grid(True, color="dimgray", axis="y")  # Plots a parallel-x-axis gird
    plt.legend()  # Displays the graph legend at the top-right most of the screen

    plt.show(block=False)  # Displays our graph and does not halt the execution of the program
    plt.pause(6)  # Pauses the execution for a specific time interval to allow the user to comprehend the graph
    plt.savefig("Figure_Class_Grades.pdf")  # initially creates an empty file
    plt.savefig("Figure_Class_Grades.pdf")  # Saves the figure as pdf


def open_excel(file_path):
    # Function to open and read data in Excel file
    from openpyxl import load_workbook
    workbook = load_workbook(file_path, read_only=True)
    return workbook


def correct_sheet_name(sheet_name, file_path):
    # Function to check whether the entered sheet name actually exists with the Excel file
    workbook = open_excel(file_path)
    sheet_names = workbook.sheetnames
    while sheet_name not in sheet_names:
        sheet_name = input("Error! this sheet name does not exist, Please enter sheet name: ")
    return sheet_name


def data_xlsx(file_path, sheet_name, column_name):
    # Function to extract the data of the Excel file
    import pandas as pd  # Imports pandas, a well-known python library for data science/analysis
    data = pd.read_excel(file_path, sheet_name=sheet_name)  # Reading midterm grades from the specified .xlsx file

    MT = []
    for index in data.index:
        each_num = data[column_name][index]
        MT.append(each_num)
    return MT


def graph_xlsx(file_path, sheet_name, column_name):
    # Function to graph the data of students into a histogram
    from matplotlib import pyplot as plt  # Imports matplotlib, a widely-known graphing and data visualization python library
    import pandas as pd  # Imports pandas, a well-known python library for data science/analysis
    import mplcyberpunk  # Imports a cool styling sheet cyberpunk-like

    data = pd.read_excel(file_path, sheet_name=sheet_name)  # Reading midterm grades from a .xlsx file
    MT_grades = data[column_name]  # Specifying the name of the column that we will graph (useful in case of many columns)

    MT_sorted = sort(MT_grades)  # To graph a histogram, numbers should be sorted and in ascending order
    Average = average(MT_sorted)  # Calculates the average of dataset
    Median = median(MT_sorted)  # Calculates the median of the dataset
    Bins = range(5, 105, 11)  # Specifying our custom bins (x-axis values)

    plt.style.use("cyberpunk")  # Sets the background color to dark instead of the default white
    plt.hist(MT_sorted, bins=Bins, edgecolor="black")  # Graphs midterm grades versus number of students in a histogram
    plt.title("Grade Distribution", fontsize=25, font="Arial")  # Sets graph title to "Grade Distribution"\

    plt.xlabel("Grades", fontsize=20)  # Sets "Grades" as the title for the x-axis
    plt.ylabel("No. Students", fontsize=20)  # Sets "No. Students" as the title for the y-axis
    plt.axvline(Median, color="red", label="Median", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our median
    plt.axvline(Average, color="orange", label="Average", linewidth=1)  # Puts an "Axis_Vertical_Line", which will represent our average
    plt.xticks(Bins)  # Displays our custom bin values instead of the default one (multiples of 10)

    plt.tight_layout()  # Positions axis titles correctly in place
    plt.grid(True, axis="y")  # Plots a parallel-x-axis gird
    plt.legend()  # Displays the graph legend at the top-right most of the screen

    plt.show(block=False)  # Displays our graph and does not halt the execution of the program
    plt.pause(5)  # Pauses the execution for a specific time interval to allow the user to comprehend the graph
    plt.savefig("Figure_Class_Grades.pdf")  # initially creates an empty file
    plt.savefig("Figure_Class_Grades.pdf")  # Saves the figure as pdf


def open_txt(file_path):
    # Function to open text file
    file = open(file_path, "r")
    return file


def data_txt(file_path):
    # Function to extract data from a text file
    file = open_txt(file_path)
    MT = []
    for line in file:
        MT.append(line)
    return MT


def MT_float(MT):
    # Function to convert from string-valued numbers of .split to integer values
    for i in range(len(MT)):
        MT[i] = float(MT[i])
    return MT


def No_attendees(MT):
    # Function to determine the number of attendees
    number_attendees = len(MT)
    return number_attendees


def average(MT):
    # Function to calculate the average of a dataset
    Sum = 0
    MT_float(MT)

    for i in range(len(MT)):
        Sum = Sum + MT[i]

    Avg = round(Sum / len(MT), 2)
    return Avg


def median(MT):
    # Function to calculate the median of a dataset
    MT_sorted = sort(MT)
    if len(MT_sorted) % 2 == 0:
        # If we have an even number of grades, the median is the average of the two numbers in the middle
        n1 = MT_sorted[int(((len(MT_sorted)) / 2) - 1)]  # We take the integer of that because be default "/" returns a float
        n2 = MT_sorted[int(((len(MT_sorted)) / 2))]  # OR WE CAN use "//" instead of "/"
        Median = (n1 + n2) / 2
    else:
        # If we have an odd number of grades, the median is the number exactly in the middle
        Median = MT_sorted[int((len(MT_sorted) - 1) / 2)]
    return Median


def standard_deviation(MT):
    # Function to calculate the standard deviation of a dataset
    import math
    Avg = average(MT)
    Sum = 0
    MT_float(MT)

    for i in range(len(MT)):
        Sum = Sum + (MT[i] - Avg)**2

    SD = round(math.sqrt(Sum / len(MT)), 2)
    return SD


def time_wait(n):
    # Functon to halt the execution of the program for a specific time interval
    import time
    time.sleep(n)




